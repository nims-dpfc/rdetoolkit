import json
import os
from pathlib import Path
import pathlib

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import pandas as pd
import pytest
from pandas.testing import assert_frame_equal
from rdetoolkit.exceptions import StructuredError
from rdetoolkit.invoicefile import (
    ExcelInvoiceFile,
    InvoiceFile,
    check_exist_rawfiles,
    read_excelinvoice,
    update_description_with_features,
    apply_magic_variable,
)
from rdetoolkit.models.rde2types import RdeOutputResourcePath
from rdetoolkit.invoicefile import ExcelInvoiceTemplateGenerator
from rdetoolkit.models.invoice import FixedHeaders, TemplateConfig


@pytest.fixture
def template_config_mode_file():
    invoice_dir = pathlib.Path("data", "tasksupport")
    invoice_schema_json_path = pathlib.Path(str(invoice_dir), "invoice.schema.json")

    static_dir = Path(__file__).parent.parent / "src" / "rdetoolkit" / "static"
    specific_term_path = static_dir / "ex_specificTerm.csv"
    general_term_path = static_dir / "ex_generalTerm.csv"
    yield TemplateConfig(
        schema_path=str(invoice_schema_json_path),
        general_term_path=general_term_path,
        specific_term_path=specific_term_path,
        inputfile_mode="file",
    )


@pytest.fixture
def template_config_mode_folder():
    invoice_dir = pathlib.Path("data", "tasksupport")
    invoice_schema_json_path = pathlib.Path(str(invoice_dir), "invoice.schema.json")

    static_dir = Path(__file__).parent.parent / "src" / "rdetoolkit" / "static"
    specific_term_path = static_dir / "ex_specificTerm.csv"
    general_term_path = static_dir / "ex_generalTerm.csv"
    yield TemplateConfig(
        schema_path=str(invoice_schema_json_path),
        general_term_path=general_term_path,
        specific_term_path=specific_term_path,
        inputfile_mode="folder",
    )


@pytest.fixture
def expected_df():
    parent_dir = Path(__file__).parent
    csv_path = parent_dir / Path("samplefile", "test_excelinvoice_term_sample.csv")
    df = pd.read_csv(csv_path, header=None)
    df.columns = df.columns.astype(str)
    yield df


@pytest.fixture
def expected_df_folder():
    parent_dir = Path(__file__).parent
    csv_path = parent_dir / Path("samplefile", "test_excelinvoice_term_sample_folder.csv")
    df = pd.read_csv(csv_path, header=None)
    df.columns = df.columns.astype(str)
    yield df


@pytest.fixture
def sample_df():
    return pd.DataFrame({
        'col1': ['a', 'b', 'c'],
        'col2': [1, 2, 3],
        'col3': ['x', 'y', 'z']
    })


def assert_frame_equal_ignore_column_names(df1, df2):
    if df1.shape != df2.shape:
        raise AssertionError(f"DataFrames have different shapes: {df1.shape} vs {df2.shape}")

    df1_temp = df1.copy()
    df2_temp = df2.copy()
    df1_temp.columns = range(df1_temp.shape[1])
    df2_temp.columns = range(df2_temp.shape[1])

    df1_temp.index.name = None
    df2_temp.index.name = None

    assert_frame_equal(df1_temp, df2_temp, check_names=False)


def test_get_item_invoice(ivnoice_json_none_sample_info):
    invoice = InvoiceFile(ivnoice_json_none_sample_info)
    assert invoice["basic"]["dateSubmitted"] == "2023-03-14"


def test_set_item_invoice(ivnoice_json_none_sample_info):
    invoice = InvoiceFile(ivnoice_json_none_sample_info)
    invoice["basic"]["dateSubmitted"] = "2023-03-15"
    assert invoice["basic"]["dateSubmitted"] == "2023-03-15"


def test_delete_item_invoice(ivnoice_json_none_sample_info):
    invoice = InvoiceFile(ivnoice_json_none_sample_info)
    del invoice["basic"]["dateSubmitted"]
    assert "dateSubmitted" not in invoice["basic"]


def test_invoicefile_read_method(ivnoice_json_none_sample_info):
    """テストケース: InvoiceFileのread"""
    invoice_file = InvoiceFile(ivnoice_json_none_sample_info)
    expect_data = {
        "datasetId": "1s1199df4-0d1v-41b0-1dea-23bf4dh09g12",
        "basic": {
            "dateSubmitted": "2023-03-14",
            "dataOwnerId": "0c233ef274f28e611de4074638b4dc43e737ab993132343532343430",
            "dataName": "test1",
            "experimentId": "test_230606_1",
            "description": "desc1",
        },
        "custom": {"key1": "test1", "key2": "test2"},
    }
    assert invoice_file.invoice_obj == expect_data


def test_overwrite_method(ivnoice_json_none_sample_info):
    """テストケース: InvoiceFileのoverwrite"""
    dst_file_path = Path("tests/test_dist_invoice.json")
    invoice_file = InvoiceFile(ivnoice_json_none_sample_info)
    invoice_file.invoice_obj["custom"]["key3"] = "test3"
    invoice_file.overwrite(dst_file_path)

    with open(ivnoice_json_none_sample_info, "rb") as src_file, open(dst_file_path, "rb") as dist_file:
        src_obj = json.load(src_file)
        dst_obj = json.load(dist_file)
        assert src_obj != dst_obj
        assert dst_obj.get("custom").get("key3") == "test3"
    if os.path.exists(dst_file_path):
        os.remove(dst_file_path)


def test_copy_method(ivnoice_json_none_sample_info):
    """テストケース: InvoiceFileのcopy"""
    dst_file_path = Path("tests/test_dist_invoice.json")
    InvoiceFile.copy_original_invoice(ivnoice_json_none_sample_info, dst_file_path)

    with open(ivnoice_json_none_sample_info, "rb") as src_file, open(dst_file_path, "rb") as dist_file:
        src_obj = json.load(src_file)
        dst_obj = json.load(dist_file)
        assert src_obj == dst_obj
    if os.path.exists(dst_file_path):
        os.remove(dst_file_path)


def test_read_valid_excel_invoice_file(inputfile_single_dummy_header_excelinvoice):
    """Excelinvoiceのheaderに適当な値が入力されても同じヘッダー情報へ変更して返すかテスト"""
    invoice_path = Path(inputfile_single_dummy_header_excelinvoice)
    excel_invoice_file = ExcelInvoiceFile(invoice_path)

    dfexcelinvoice, df_general, df_specific = excel_invoice_file.read()

    assert dfexcelinvoice.columns[0] == "data_file_names/name"
    assert (df_general.columns == ["term_id", "key_name"]).all()
    assert (df_specific.columns == ["sample_class_id", "term_id", "key_name"]).all()


def test_read_none_sample_excel_invoice_file(excelinvoice_non_sampleinfo):
    """Excelinvoiceのsample情報なしの時のreadテスト"""
    invoice_path = Path(excelinvoice_non_sampleinfo)
    excel_invoice_file = ExcelInvoiceFile(invoice_path)

    dfexcelinvoice, df_general, df_specific = excel_invoice_file.read()

    assert dfexcelinvoice.columns[0] == "data_file_names/name"
    assert (df_general.columns == ["term_id", "key_name"]).all()
    assert (df_specific.columns == ["sample_class_id", "term_id", "key_name"]).all()


def test_read_none_sample_excel_invoice_file_blankline(
    inputfile_single_excelinvoice_with_blankline,
):
    """Excelinvoiceの行間に空行がある場合のテスト
    想定としては、エクセルインボイスの5行目移行に空行がある場合エラーメッセージを出す
    """
    invoice_path = Path(inputfile_single_excelinvoice_with_blankline)

    with pytest.raises(StructuredError) as e:
        excel_invoice_file = ExcelInvoiceFile(invoice_path)
        dfexcelinvoice, df_general, df_specific = excel_invoice_file.read()

    assert str(e.value) == "Error! Blank lines exist between lines"


def test_excelinvoice_overwrite(inputfile_multi_excelinvoice, ivnoice_json_with_sample_info, ivnoice_schema_json_none_specificAttributes):
    """試料情報ありの上書き処理
    上書き後のinvoice.jsonの内容を確認する
    上書き前のkey/value -> custom.key1: test1
    上書き前のkey/value -> custom.key2: test2
    """
    dist_path = Path("data", "invoice", "invoice.json")
    excel_invoice_path = Path(inputfile_multi_excelinvoice)

    excel_invoice_file = ExcelInvoiceFile(excel_invoice_path)
    excel_invoice_file.overwrite(ivnoice_json_with_sample_info, dist_path, ivnoice_schema_json_none_specificAttributes, 0)

    with open(dist_path, encoding="utf-8") as f:
        contents = json.load(f)

    assert contents["custom"]["key1"] == "AAA"
    assert contents["custom"]["key2"] == "CCC"


def test_excelinvoice_overwrite_none_sample(excelinvoice_non_sampleinfo, ivnoice_json_none_sample_info, ivnoice_schema_json_none_sample):
    """試料情報なしの上書き処理
    上書き後のinvoice.jsonの内容を確認する
    上書き前のkey/value -> custom.key1: test1
    上書き前のkey/value -> custom.key2: test2
    """
    dist_path = Path("data", "invoice", "invoice.json")
    excel_invoice_path = Path(excelinvoice_non_sampleinfo)

    excel_invoice_file = ExcelInvoiceFile(excel_invoice_path)
    excel_invoice_file.overwrite(ivnoice_json_none_sample_info, dist_path, ivnoice_schema_json_none_sample, 0)

    with open(dist_path, encoding="utf-8") as f:
        contents = json.load(f)

    assert contents["custom"]["key1"] == "AAA"
    assert contents["custom"]["key2"] == "CCC"


def test_read_invalid_excel_invoice_file():
    invoice_path = Path("dummy/file.xlsx")
    with pytest.raises(StructuredError) as e:
        ExcelInvoiceFile(invoice_path)
    assert str(e.value) == "ERROR: excelinvoice not found dummy/file.xlsx"


def test_read_no_first_sheet_excel_invoice_file(inputfile_empty_excelinvoice):
    invoice_path = Path(inputfile_empty_excelinvoice)
    with pytest.raises(StructuredError) as e:
        ExcelInvoiceFile(invoice_path)
    assert str(e.value) == "ERROR: no sheet in invoiceList files"


@pytest.fixture
def rde_resource():
    rde_resource = RdeOutputResourcePath(
        raw=Path("dummy"),
        rawfiles=tuple(["dummy1", "dummy2"]),
        struct=Path("dummmy"),
        main_image=Path("dummmy"),
        other_image=Path("dummmy"),
        thumbnail=Path("dummmy"),
        meta=Path("data/meta"),
        logs=Path("dummmy"),
        invoice=Path("data/invoice"),
        invoice_schema_json=Path("data/tasksupport/invoice.schema.json"),
        invoice_org=Path("data/invoice/invoice.json"),
        nonshared_raw=Path("dummmy"),
    )
    yield rde_resource


def test_update_description_with_features(
    rde_resource,
    ivnoice_schema_json,
    metadata_def_json_with_feature,
    ivnoice_json_none_sample_info,
    metadata_json,
):
    """テストケース: descriptionへの書き出しがパスするかテスト"""
    expect_message = "desc1\n特徴量1:test-value1\n特徴量2(V):test-value2\n特徴量3(V):test-value3"

    # 検証
    update_description_with_features(rde_resource, ivnoice_json_none_sample_info, metadata_def_json_with_feature)

    # 書き込み結果を比較
    with open(ivnoice_json_none_sample_info, encoding="utf-8") as f:
        result_contents = json.load(f)

    assert result_contents["basic"]["description"] == expect_message


def test_update_description_with_features_missing_target_key(
    rde_resource,
    ivnoice_schema_json,
    metadata_def_json_with_feature,
    ivnoice_json_none_sample_info,
    metadata_json_missing_value,
):
    """テストケース:
    descriptionへの書き出すはずのメタデータのvalueが存在しないとき、descriptionが記述できるかどうかテスト
    想定としては、書き出す対象のメタデータのvalueがなくても記述できることを想定している。
    """
    expect_message = "desc1\n特徴量1:test-value1\n特徴量3(V):test-value3"

    # 検証
    update_description_with_features(rde_resource, ivnoice_json_none_sample_info, metadata_def_json_with_feature)

    # 書き込み結果を比較
    with open(ivnoice_json_none_sample_info, encoding="utf-8") as f:
        result_contents = json.load(f)

    assert result_contents["basic"]["description"] == expect_message


def test_update_description_with_features_missing_target_key_none_result_metadata(
    rde_resource,
    ivnoice_schema_json,
    metadata_def_json_with_feature,
    ivnoice_json_none_sample_info,
    metadata_json_missing_value,
):
    """テストケース:
    metadata.jsonにconstant, variableがない場合でも、descriptionの処理を正しく実行できるか確認。
    constant, variableがない場合、descriptionにコメントを追加せずに処理をパスする。
    """
    expect_message = "desc1\n特徴量1:test-value1\n特徴量3(V):test-value3"

    # 検証
    update_description_with_features(rde_resource, ivnoice_json_none_sample_info, metadata_def_json_with_feature)

    # 書き込み結果を比較
    with open(ivnoice_json_none_sample_info, encoding="utf-8") as f:
        result_contents = json.load(f)

    assert result_contents["basic"]["description"] == expect_message


def test_update_description_with_features_none_variable(
    rde_resource,
    ivnoice_schema_json,
    metadata_def_json_with_feature,
    ivnoice_json_none_sample_info,
    metadata_json_non_variable,
):
    """テストケース: metadata.jsonのconstantに全て特徴量の情報が記述されており、
    descriptionへの書き出しがパスするかテスト
    """
    expect_message = "desc1\n特徴量1:test-value1\n特徴量2(V):test-value2\n特徴量3(V):test-value3"

    # 検証
    update_description_with_features(rde_resource, ivnoice_json_none_sample_info, metadata_def_json_with_feature)

    # 書き込み結果を比較
    with open(ivnoice_json_none_sample_info, encoding="utf-8") as f:
        result_contents = json.load(f)

    assert result_contents["basic"]["description"] == expect_message


def test_update_description_with_features_none_constant(
    rde_resource,
    ivnoice_schema_json,
    metadata_def_json_with_feature,
    ivnoice_json_none_sample_info,
    metadata_json_non_constat,
):
    """テストケース: metadata.jsonのconstantに全て特徴量の情報が記述されており、
    descriptionへの書き出しがパスするかテスト
    """
    expect_message = "desc1\n特徴量1:test-value1\n特徴量2(V):test-value2\n特徴量3(V):test-value3"

    # 検証
    update_description_with_features(rde_resource, ivnoice_json_none_sample_info, metadata_def_json_with_feature)

    # 書き込み結果を比較
    with open(ivnoice_json_none_sample_info, encoding="utf-8") as f:
        result_contents = json.load(f)

    assert result_contents["basic"]["description"] == expect_message


def test_update_description_none_features_none_variable(
    rde_resource,
    ivnoice_schema_json,
    metadata_def_json_none_feature,
    ivnoice_json_none_sample_info,
    metadata_json_non_variable,
):
    """テストケース:
    metadata.jsonのconstantにメタデータの記載はあるが、featureがないmetadata.josnを正しくPassできるかテスト
    """
    expect_message = "desc1"

    # 検証
    update_description_with_features(rde_resource, ivnoice_json_none_sample_info, metadata_def_json_none_feature)
    # 書き込み結果を比較
    with open(ivnoice_json_none_sample_info, encoding="utf-8") as f:
        result_contents = json.load(f)

    assert result_contents["basic"]["description"] == expect_message


def test_read_excelinvoice(inputfile_single_excelinvoice):
    """read_excelinvoiceのテスト
    dfexcelinvoice, df_general, dfSpecificが正しい値で返ってくるかテスト
    また、空のシートが含まれるエクセルインボイスを入れた時に想定通りの値を出力するかテスト
    """
    expect_sheet1 = [
        [
            "test_child1.txt",
            "N_TEST_1",
            "test_user",
            "de17c7b3f0ff5126831c2d519f481055ba466ddb6238666132316439",
            "test1",
            "ee17c7b3-f0ff-5126-831c-2d519f481055",
            "test_230606_1",
            "https://sample.com",
            "desc1",
            "sample1",
            "de17c7b3-f0ff-5126-831c-2d519f481055",
            "de17c7b3f0ff5126831c2d519f481055ba466ddb6238666132316439",
            "sample1",
            "https://sample.com",
            "desc3",
            "testname",
            "Fe",
            "magnet",
            "7439-89-6",
            "7439-89-6",
            "7439-89-6",
            "7439-89-6",
            "AAA",
            "CCC",
        ],
    ]
    df1 = pd.DataFrame(
        expect_sheet1,
        columns=[
            "data_file_names/name",
            "dataset_title",
            "dataOwner",
            "basic/dataOwnerId",
            "basic/dataName",
            "basic/instrumentId",
            "basic/experimentId",
            "basic/referenceUrl",
            "basic/description",
            "sample/names",
            "sample/sampleId",
            "sample/ownerId",
            "sample/composition",
            "sample/referenceUrl",
            "sample/description",
            "sample.general/general-name",
            "sample.general/cas-number",
            "sample.general/crystal-structure",
            "sample.general/purchase-date",
            "sample.general/lot-number-or-product-number-etc",
            "sample.general/smiles-string",
            "sample.general/supplier",
            "custom/key1",
            "custom/key2",
        ],
    )

    dfexcelinvoice, df_general, df_specific = read_excelinvoice(inputfile_single_excelinvoice)

    assert_frame_equal(dfexcelinvoice, df1)
    assert isinstance(df_general, pd.DataFrame)
    assert df_general.columns.to_list() == ["term_id", "key_name"]
    assert isinstance(df_specific, pd.DataFrame)
    assert df_specific.columns.to_list() == ["sample_class_id", "term_id", "key_name"]


def test_empty_excelinvoice_read_excelinvoice(empty_inputfile_excelinvoice):
    """空のエクセルインボイスを入れた時に例外をキャッチできるかテスト"""
    with pytest.raises(StructuredError) as e:
        _, _, _ = read_excelinvoice(empty_inputfile_excelinvoice)
    assert str(e.value) == "ERROR: no sheet in invoiceList files"


def test_invalid_excelinvoice_read_excelinvoice(
    inputfile_invalid_samesheet_excelinvoice,
):
    """sheet1の内容が複数あるエクセルインボイスを入れた時に例外をキャッチできるかテスト"""
    with pytest.raises(StructuredError) as e:
        _, _, _ = read_excelinvoice(inputfile_invalid_samesheet_excelinvoice)
    assert str(e.value) == "ERROR: multiple sheet in invoiceList files"


def test_check_exist_rawfiles(inputfile_multi_excelinvoice):
    """excelinvoiceに記述されたファイル名通りにファイルの並び替えが実行されるかテスト

    inputfile_multi_excelinvoiceには、テスト用のエクセルインボイスのパスが格納
    """
    expect_rtn = [Path("test_child1.txt"), Path("test_child2.txt")]

    # テスト用エクセルインボイス前処理
    df_excelinvoice = pd.read_excel(inputfile_multi_excelinvoice, sheet_name="invoice_form", dtype=str, header=None, index_col=None)
    df = df_excelinvoice.dropna(axis=0, how="all").dropna(axis=1, how="all")
    hd1 = list(df.iloc[1, :].fillna(""))
    hd2 = list(df.iloc[2, :].fillna(""))
    df.columns = [f"{s1}/{s2}" if s1 else s2 for s1, s2 in zip(hd1, hd2)]
    df_excelinvoice = df.iloc[4:, :].reset_index(drop=True).copy()

    test_excel_raw_files = [
        Path("test_child2.txt"),
        Path("test_child1.txt"),
        Path("test_child3.txt"),
        Path("test_child9.txt"),
        Path("test_child10.txt"),
    ]
    rtn = check_exist_rawfiles(df_excelinvoice, test_excel_raw_files)

    assert expect_rtn == rtn


def test_error_check_exist_rawfiles(inputfile_multi_excelinvoice):
    """excelinvoiceに記述されたファイル名がtempフォルダリストに格納されていないとき、エラーが正しく出力されるかテスト

    inputfile_multi_excelinvoiceには、テスト用のエクセルインボイスのパスが格納
    """
    # テスト用エクセルインボイス前処理
    df_excelinvoice = pd.read_excel(inputfile_multi_excelinvoice, sheet_name="invoice_form", dtype=str, header=None, index_col=None)
    df = df_excelinvoice.dropna(axis=0, how="all").dropna(axis=1, how="all")
    hd1 = list(df.iloc[1, :].fillna(""))
    hd2 = list(df.iloc[2, :].fillna(""))
    df.columns = [f"{s1}/{s2}" if s1 else s2 for s1, s2 in zip(hd1, hd2)]
    df_excelinvoice = df.iloc[4:, :].reset_index(drop=True).copy()

    # テスト用rawファイル群
    test_excel_raw_files = [Path("test_child1.txt")]

    with pytest.raises(StructuredError) as e:
        _ = check_exist_rawfiles(df_excelinvoice, test_excel_raw_files)

    assert str(e.value) == "ERROR: raw file not found: test_child2.txt"


def test_apply_magic_variable_inputfile_check(ivnoice_json_magic_filename_variable):
    """${filename}が置換できるかテスト

    conftest.pyよりfixsture: ivnoice_json_magic_filename_variableを取得
    """
    invoice_path = ivnoice_json_magic_filename_variable
    rawfile_path = "/test/dummy/dymmy_replace_filename.txt"

    result = apply_magic_variable(invoice_path, rawfile_path)

    assert result["basic"]["dataName"] == "dymmy_replace_filename.txt"


class TestExcelinvoice:
    """Excelinvoiceクラスのテスト"""

    def test_read(self, inputfile_single_excelinvoice):
        """read_excelinvoiceのテスト
        dfexcelinvoice, df_general, dfSpecificが正しい値で返ってくるかテスト
        また、空のシートが含まれるエクセルインボイスを入れた時に想定通りの値を出力するかテスト
        """
        expect_sheet1 = [
            [
                "test_child1.txt",
                "N_TEST_1",
                "test_user",
                "de17c7b3f0ff5126831c2d519f481055ba466ddb6238666132316439",
                "test1",
                "ee17c7b3-f0ff-5126-831c-2d519f481055",
                "test_230606_1",
                "https://sample.com",
                "desc1",
                "sample1",
                "de17c7b3-f0ff-5126-831c-2d519f481055",
                "de17c7b3f0ff5126831c2d519f481055ba466ddb6238666132316439",
                "sample1",
                "https://sample.com",
                "desc3",
                "testname",
                "Fe",
                "magnet",
                "7439-89-6",
                "7439-89-6",
                "7439-89-6",
                "7439-89-6",
                "AAA",
                "CCC",
            ],
        ]
        df1 = pd.DataFrame(
            expect_sheet1,
            columns=[
                "data_file_names/name",
                "dataset_title",
                "dataOwner",
                "basic/dataOwnerId",
                "basic/dataName",
                "basic/instrumentId",
                "basic/experimentId",
                "basic/referenceUrl",
                "basic/description",
                "sample/names",
                "sample/sampleId",
                "sample/ownerId",
                "sample/composition",
                "sample/referenceUrl",
                "sample/description",
                "sample.general/general-name",
                "sample.general/cas-number",
                "sample.general/crystal-structure",
                "sample.general/purchase-date",
                "sample.general/lot-number-or-product-number-etc",
                "sample.general/smiles-string",
                "sample.general/supplier",
                "custom/key1",
                "custom/key2",
            ],
        )

        excelinvoice = ExcelInvoiceFile(inputfile_single_excelinvoice)
        dfexcelinvoice, df_general, df_specific = excelinvoice.read()
        assert_frame_equal(dfexcelinvoice, df1)
        assert isinstance(df_general, pd.DataFrame)
        assert df_general.columns.to_list() == ["term_id", "key_name"]
        assert isinstance(df_specific, pd.DataFrame)
        assert df_specific.columns.to_list() == ["sample_class_id", "term_id", "key_name"]

    def test_read_emptyfile(self, empty_inputfile_excelinvoice):
        """空のエクセルインボイスを入れた時に例外をキャッチできるかテスト"""
        with pytest.raises(StructuredError) as e:
            excelinvoice = ExcelInvoiceFile(empty_inputfile_excelinvoice)
            _, _, _ = excelinvoice.read()
        assert str(e.value) == "ERROR: no sheet in invoiceList files"

    def test_invalidfile_read(self, inputfile_invalid_samesheet_excelinvoice):
        """sheet1の内容が複数あるエクセルインボイスを入れた時に例外をキャッチできるかテスト"""
        with pytest.raises(StructuredError) as e:
            excelinvoice = ExcelInvoiceFile(inputfile_invalid_samesheet_excelinvoice)
            _, _, _ = excelinvoice.read()
        assert str(e.value) == "ERROR: multiple sheet in invoiceList files"

    def test_generate_template(self, inputfile_single_excelinvoice, ivnoice_schema_json_with_full_sample_info, expected_df):
        excelinvoice = ExcelInvoiceFile(inputfile_single_excelinvoice)
        template_df = excelinvoice.generate_template(ivnoice_schema_json_with_full_sample_info, save_path="test_excelinvoice.xlsx")

        assert template_df.iloc[0, 0] == "invoiceList_format_id"
        assert os.path.exists("test_excelinvoice.xlsx")

        if os.path.exists("test_excelinvoice.xlsx"):
            os.remove("test_excelinvoice.xlsx")

    def test_save_basic(self, inputfile_single_excelinvoice):
        excelinvoice = ExcelInvoiceFile(inputfile_single_excelinvoice)
        excelinvoice.save("test_excelinvoice.xlsx")
        assert os.path.exists("test_excelinvoice.xlsx")

    def test_save_with_custom_params(self, inputfile_single_excelinvoice, sample_df):
        sheet_name = "カスタムシート"
        index = ['列1', '列2', '列3']
        header = ['行1', '行2', '行3']
        excelinvoice = ExcelInvoiceFile(inputfile_single_excelinvoice)
        excelinvoice.save(
            "test_excelinvoice.xlsx",
            invoice=sample_df,
            index=index,
            header=header,
            sheet_name=sheet_name
        )
        assert os.path.exists("test_excelinvoice.xlsx")

        saved_df = pd.read_excel("test_excelinvoice.xlsx")
        assert saved_df.iloc[:, 0].to_list() == index
        assert saved_df.columns.to_list()[1:] == header
        with pd.ExcelFile("test_excelinvoice.xlsx") as xl:
            assert sheet_name in xl.sheet_names

        if os.path.exists("test_excelinvoice.xlsx"):
            os.remove("test_excelinvoice.xlsx")

    def test_save_failed(self, inputfile_single_excelinvoice, sample_df):
        sheet_name = "カスタムシート"
        header = ['列1', '列2', '列3']
        index = ['行1', '行2']
        with pytest.raises(StructuredError) as e:
            excelinvoice = ExcelInvoiceFile(inputfile_single_excelinvoice)
            excelinvoice.save(
                "test_excelinvoice.xlsx",
                invoice=sample_df,
                index=index,
                header=header,
                sheet_name=sheet_name
            )
            assert str(e.value) == "Failed to save the invoice file."

        if os.path.exists("test_excelinvoice.xlsx"):
            os.remove("test_excelinvoice.xlsx")


def test_generate_basic_filemode(template_config_mode_file, ivnoice_schema_json_with_full_sample_info, inputfile_single_excelinvoice, expected_df):
    generator = ExcelInvoiceTemplateGenerator(FixedHeaders())
    result_df = generator.generate(template_config_mode_file)
    assert_frame_equal_ignore_column_names(result_df, expected_df)


def test_generate_basic_foldermode(template_config_mode_folder, ivnoice_schema_json_with_full_sample_info, inputfile_single_excelinvoice, expected_df_folder):
    generator = ExcelInvoiceTemplateGenerator(FixedHeaders())
    result_df = generator.generate(template_config_mode_folder)
    print(result_df.head(10))
    assert result_df.iloc[1, 0] == ''
    assert result_df.iloc[2, 0] == "data_folder"


def test_save_basic_file_creation(template_config_mode_file, ivnoice_schema_json_with_full_sample_info, inputfile_single_excelinvoice, expected_df):
    test_path = "test_excelinvoice.xlsx"
    generator = ExcelInvoiceTemplateGenerator(FixedHeaders())
    generator.save(expected_df, test_path)

    assert os.path.exists(test_path)

    wb = load_workbook(test_path)
    ws = wb['invoice_form']
    assert ws.row_dimensions[4].height == 40

    for col in range(1, expected_df.shape[1] + 1):
        assert ws.column_dimensions[get_column_letter(col)].width == 20

    if os.path.exists(test_path):
        os.remove(test_path)


def test_save_border_settings(template_config_mode_file, ivnoice_schema_json_with_full_sample_info, inputfile_single_excelinvoice, expected_df):
    test_path = "test_excelinvoice.xlsx"
    generator = ExcelInvoiceTemplateGenerator(FixedHeaders())
    generator.save(expected_df, test_path)

    wb = load_workbook(test_path)
    ws = wb['invoice_form']
    for col in range(1, expected_df.shape[1] + 1):
        cell = ws.cell(row=4, column=col)
        assert cell.border.top.style == "thick"
        assert cell.border.bottom.style == "double"

    # グリッド罫線チェック
    for row in range(5, 41):
        for col in range(1, expected_df.shape[1] + 1):
            cell = ws.cell(row=row, column=col)
            assert cell.border.top.style == "thin"
            assert cell.border.bottom.style == "thin"
            assert cell.border.left.style == "thin"
            assert cell.border.right.style == "thin"

    if os.path.exists(test_path):
        os.remove(test_path)
